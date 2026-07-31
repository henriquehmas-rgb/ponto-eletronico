"""Exportador XLSX, streaming linha a linha (F11, T5/A3).

`openpyxl` ja e dependencia do worker (F2, leitura de planilha de
importacao); a API ganhou a mesma biblioteca nesta fase (`apps/api/
pyproject.toml`, bloco `# --- F11 ---`) porque `executarRelatorio` responde
`200` com o arquivo direto na API para relatorios leves em `formato=xlsx`
(PCF §2.5).

`Workbook(write_only=True)` e o modo de escrita otimizada do proprio
`openpyxl`: cada `worksheet.append(...)` grava a linha direto no arquivo
(zip) em disco/buffer, sem manter a planilha inteira em memoria -- e o que
prova o criterio de aceite T5 ("exportar 10.000 linhas em XLSX nao estoura
memoria").
"""

from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from typing import Any, BinaryIO

import openpyxl

from app.relatorios.exportadores import ColunaExportacao, cabecalhos, valor_formatado

#: Nome da unica aba do arquivo exportado -- generico de proposito, o nome do
#: relatorio ja vai no cabecalho da requisicao/nome do arquivo, nao precisa
#: duplicar aqui.
NOME_ABA = "Relatorio"


def exportar(
    linhas: Iterable[Mapping[str, Any]],
    colunas: Sequence[ColunaExportacao],
    *,
    destino: BinaryIO,
    converter_decimal: bool = False,
    casas_decimais: int = 2,
) -> None:
    """Escreve `linhas` como XLSX em `destino`, em modo `write_only` (streaming,
    sem materializar a planilha inteira em memoria -- PCF §2.3 item 2).
    `destino` permanece aberto ao final."""
    pasta = openpyxl.Workbook(write_only=True)
    aba = pasta.create_sheet(title=NOME_ABA)
    aba.append(cabecalhos(colunas))
    for linha in linhas:
        aba.append(
            [
                valor_formatado(
                    linha,
                    coluna,
                    converter_decimal=converter_decimal,
                    casas_decimais=casas_decimais,
                )
                for coluna in colunas
            ]
        )
    pasta.save(destino)


__all__ = ["NOME_ABA", "exportar"]
