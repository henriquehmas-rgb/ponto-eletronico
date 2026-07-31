"""T15 -- Testes de propriedade e e2e do catálogo COMPLETO de 24 relatórios
(PCF F11 §6/T15, critério de aceite oficial 1, tarefa CONJUNTA A1/A2/A3/A4,
fechada no T16 -- nenhum dos quatro agentes rodou isto individualmente).

Para cada um dos 24 códigos semeados por `tests/f11/conftest.py::
_CATALOGO_RELATORIOS` (a MESMA lista que ancora a fixture -- importada daqui,
não duplicada, para nunca divergir de quais 24 códigos realmente existem):

1. Executa o relatório (síncrono quando `RelatorioDefinicao.assincrono` é
   falso; assíncrono -- chamando `worker.tarefas.relatorios.executar_
   relatorio` diretamente, mesmo padrão de `tests/f11/agendamentos/
   test_worker_executar_relatorio.py`, em vez de esperar um worker real
   consumir a fila -- quando é verdadeiro), para cada formato declarado em
   `RelatorioDefinicao.formatos`.
2. Confere que a execução conclui (`status='concluido'`, nunca
   `'falhou'`/preso em `'processando'`).
3. Baixa o artefato do MinIO (`app.comum.armazenamento.obter_objeto`) e prova
   que não está vazio.
4. Abre o artefato sem erro no formato certo: `csv` (módulo `csv` nativo,
   confirma pelo menos a linha de cabeçalho), `xlsx` (`openpyxl.
   load_workbook`, leitura), `pdf` (`pypdf.PdfReader`, confirma pelo menos
   1 página) -- dependências de TESTE apenas (leitura), nunca usadas pelo
   código de produção para validar o que ele mesmo escreveu.

**Exceção documentada (item 1, `espelho-oficial`):** só exporta `pdf`
(`RelatorioDefinicao.formatos == ["pdf"]` na fixture) -- não faz sentido
exportar CSV/XLSX para uma listagem de documentos já assinados/oficiais
(PCF §6/T15: "com a ressalva documentada de que o item 1 (espelho oficial)
só exporta PDF"). Este teste não força os outros dois formatos para esse
código; o parametrize abaixo usa exatamente `definicao.formatos` (lido do
banco, não uma lista fixa), então a exceção é automática -- e explicitada de
novo no `assert` dedicado no fim do módulo.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import io
import uuid
from collections.abc import AsyncIterator

import openpyxl
import pytest
import pytest_asyncio
from ponto_contracts import AssinaturaEspelho, Espelho, RelatorioDefinicao, RelatorioExecucao
from pypdf import PdfReader
from sqlalchemy.ext.asyncio import AsyncSession

# Efeito colateral: registra os 24 datasets reais (A2/A3/A4) em
# `app.relatorios.catalogo` -- sem isto, `obter_dataset` responde
# `PONTO-INT-005` para todo codigo (mesmo padrão de
# `app/routers/relatorios.py`/`tests/f11/performance/
# test_performance_relatorio.py`).
from app.relatorios import datasets as _datasets_registrados  # noqa: F401
from app.relatorios import execucao as execucao_servico
from tests.f11.conftest import _CATALOGO_RELATORIOS, ContextoF11, aplicar_tenant_teste

#: Mesma lista que ancora a fixture (T1, A1) -- única fonte de verdade de
#: quais 24 códigos existem. Importar em vez de duplicar garante que este
#: teste nunca fica desalinhado se a fixture ganhar/perder um código.
_CODIGOS_CATALOGO = [item.codigo for item in _CATALOGO_RELATORIOS]

assert len(_CODIGOS_CATALOGO) == 24, "O catalogo desta fase tem que ter exatamente 24 relatorios."


@pytest_asyncio.fixture(autouse=True)
async def _descartar_engine_global_apos_teste() -> AsyncIterator[None]:
    """Mesmo achado de infraestrutura documentado em `tests/f11/agendamentos/
    test_worker_executar_relatorio.py`: `worker.tarefas.relatorios.
    executar_relatorio` abre a própria sessão pela engine GLOBAL do processo
    (`app.db.sessao.fabrica_de_sessoes()`), diferente da engine por-teste que
    `sessao_f11` usa -- no Windows, uma conexão asyncpg aberta num loop de
    evento e finalizada noutro quebra o `pool_pre_ping`. Descartar a engine
    global ao fim de cada teste força recriação limpa no loop seguinte."""
    yield
    from app.db.sessao import encerrar_engine

    await encerrar_engine()


async def _executar_sincrono(
    sessao: AsyncSession, contexto: ContextoF11, *, codigo: str, formato: str
) -> RelatorioExecucao:
    parametros = execucao_servico.ParametrosExecucao(
        formato=formato,
        periodo_id=None,
        de=None,
        ate=None,
        empresa_id=None,
        unidade_id=None,
        departamento_id=None,
        colaborador_id=None,
        agrupamento=None,
        colunas=None,
        filtros=None,
        converter_decimal=None,
        incluir_inativos=None,
        assincrono=None,
    )
    return await execucao_servico.executar_relatorio_pedido(
        sessao,
        contexto.tenant_id,
        codigo,
        parametros,
        usuario_id=contexto.usuario_rh_id,
        redis_url="redis://unused-nao-deve-ser-chamado-no-caminho-sincrono/0",
    )


async def _executar_assincrono_direto(
    sessao: AsyncSession, contexto: ContextoF11, *, codigo: str, formato: str
) -> RelatorioExecucao:
    """Cria a linha `RelatorioExecucao(status='enfileirado')` e chama
    `worker.tarefas.relatorios.executar_relatorio` diretamente -- mesmo
    padrão de `test_worker_executar_relatorio.py` -- em vez de enfileirar de
    verdade e esperar um worker real consumir (nenhum worker roda durante os
    testes desta suite)."""
    from worker.tarefas.relatorios import executar_relatorio

    definicao_id = contexto.relatorio_ids[codigo]
    execucao_id = uuid.uuid4()
    execucao = RelatorioExecucao(
        id=execucao_id,
        tenant_id=contexto.tenant_id,
        relatorio_definicao_id=definicao_id,
        usuario_id=contexto.usuario_rh_id,
        formato=formato,
        status="enfileirado",
        progresso=0,
    )
    sessao.add(execucao)
    await sessao.commit()
    await aplicar_tenant_teste(sessao, contexto.tenant_id)

    resultado = await executar_relatorio(
        {"job_id": f"teste-e2e-{codigo}-{formato}"},
        tenant_id=str(contexto.tenant_id),
        execucao_id=str(execucao_id),
        codigo=codigo,
        parametros={},
        formato=formato,
    )
    assert resultado["sucesso"] is True, (
        f"worker.tarefas.relatorios.executar_relatorio falhou para "
        f"codigo={codigo!r} formato={formato!r}: {resultado}"
    )

    recarregada = await sessao.get(RelatorioExecucao, execucao_id, populate_existing=True)
    assert recarregada is not None
    return recarregada


def _abre_sem_erro(formato: str, conteudo: bytes, *, codigo: str) -> None:
    assert len(conteudo) > 0, f"artefato vazio para codigo={codigo!r} formato={formato!r}"
    if formato == "csv":
        texto = conteudo.decode("utf-8-sig")
        linhas = list(csv.reader(io.StringIO(texto)))
        assert len(linhas) >= 1, f"CSV sem nenhuma linha (nem cabecalho) para {codigo!r}"
    elif formato == "xlsx":
        pasta = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True)
        assert pasta.active is not None
        assert len(pasta.sheetnames) >= 1
        pasta.close()
    elif formato == "pdf":
        leitor = PdfReader(io.BytesIO(conteudo))
        assert len(leitor.pages) >= 1, f"PDF sem nenhuma pagina para {codigo!r}"
    else:  # pragma: no cover - os 24 codigos so declaram csv/xlsx/pdf
        pytest.fail(f"formato inesperado no catalogo: {formato!r}")


@pytest_asyncio.fixture
async def _espelho_oficial_de_exemplo(sessao_f11: AsyncSession, contexto_f11: ContextoF11) -> None:
    """Semeia um `Espelho`/`AssinaturaEspelho` de exemplo, localmente (regra
    de `tests/f11/conftest.py`: "cada teste que precisar de dado adicional
    cria a própria linha extra localmente") -- só para o dataset
    `espelho-oficial` (item 1) ter pelo menos uma linha real no artefato
    exportado, em vez de testar só o caso trivial de zero linhas (que também
    seria válido: `espelho_oficial` nunca gera espelho novo, PCF §6/T12)."""
    colaborador = contexto_f11.colaboradores[0]
    espelho = Espelho(
        tenant_id=contexto_f11.tenant_id,
        periodo_id=contexto_f11.periodo_id,
        colaborador_id=colaborador.colaborador_id,
        vinculo_id=colaborador.vinculo_id,
        versao=1,
        tipo="oficial",
        conteudo={"dias": []},
        hash_sha256=hashlib.sha256(f"e2e-{colaborador.colaborador_id}".encode()).hexdigest(),
    )
    sessao_f11.add(espelho)
    await sessao_f11.flush()
    sessao_f11.add(
        AssinaturaEspelho(
            tenant_id=contexto_f11.tenant_id,
            espelho_id=espelho.id,
            signatario_tipo="colaborador",
            hash_assinado=espelho.hash_sha256,
            status="assinado",
            carimbo_tempo=dt.datetime.now(tz=dt.UTC),
        )
    )
    await sessao_f11.commit()
    await aplicar_tenant_teste(sessao_f11, contexto_f11.tenant_id)


@pytest.mark.parametrize("codigo", _CODIGOS_CATALOGO)
async def test_relatorio_gera_e_exporta_nos_formatos_declarados(
    codigo: str,
    sessao_f11: AsyncSession,
    contexto_f11: ContextoF11,
    _espelho_oficial_de_exemplo: None,
) -> None:
    """Critério de aceite 1 (PCF §7): os 24 relatórios do catálogo geram e
    exportam nos formatos declarados -- síncrono quando pequeno, assíncrono
    quando `RelatorioDefinicao.assincrono=true`."""
    definicao = await sessao_f11.get(RelatorioDefinicao, contexto_f11.relatorio_ids[codigo])
    assert definicao is not None
    assert definicao.codigo == codigo

    formatos = list(definicao.formatos or [])
    assert formatos, f"relatorio {codigo!r} nao declara nenhum formato exportavel"

    if codigo == "espelho-oficial":
        assert formatos == ["pdf"], (
            "excecao documentada (PCF §6/T15): o espelho de ponto oficial so "
            "exporta PDF -- nao faz sentido CSV/XLSX para uma listagem de "
            "documentos ja assinados/oficiais."
        )

    for formato in formatos:
        if definicao.assincrono:
            execucao = await _executar_assincrono_direto(
                sessao_f11, contexto_f11, codigo=codigo, formato=formato
            )
        else:
            execucao = await _executar_sincrono(
                sessao_f11, contexto_f11, codigo=codigo, formato=formato
            )

        assert execucao.status == "concluido", (
            f"codigo={codigo!r} formato={formato!r} terminou com status="
            f"{execucao.status!r} (erro={getattr(execucao, 'erro', None)!r})"
        )
        assert execucao.conteudo_ref is not None
        assert execucao.hash_sha256 is not None

        from app.comum.armazenamento import obter_objeto

        conteudo = await obter_objeto(execucao.conteudo_ref)
        assert hashlib.sha256(conteudo).hexdigest() == execucao.hash_sha256
        _abre_sem_erro(formato, conteudo, codigo=codigo)


def test_excecao_espelho_oficial_documentada_no_proprio_catalogo() -> None:
    """Prova, sem precisar de banco, que a exceção do item 1 (só PDF) está
    fixada na PRÓPRIA semente do catálogo (`tests/f11/conftest.py`) -- não é
    um acidente deste teste específico, é a definição de fábrica."""
    (item,) = (i for i in _CATALOGO_RELATORIOS if i.codigo == "espelho-oficial")
    assert item.formatos == ["pdf"]
    outros = [i for i in _CATALOGO_RELATORIOS if i.codigo != "espelho-oficial"]
    assert all(set(i.formatos) == {"csv", "xlsx", "pdf"} for i in outros), (
        "todo relatorio alem do espelho oficial declara os tres formatos "
        "(csv/xlsx/pdf) -- a excecao e SO do item 1."
    )
