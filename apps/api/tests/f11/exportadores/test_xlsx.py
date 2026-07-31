"""`app.relatorios.exportadores.xlsx` (F11, T5/A3)."""

from __future__ import annotations

import io
import tracemalloc
from collections.abc import Iterator
from typing import Any

import openpyxl

from app.relatorios.exportadores import ColunaExportacao
from app.relatorios.exportadores import xlsx as exportador_xlsx

_COLUNAS = [
    ColunaExportacao(chave="nome", rotulo="Nome"),
    ColunaExportacao(chave="minutos", rotulo="Horas", duracao=True),
]


def _linhas() -> list[dict[str, Any]]:
    return [
        {"nome": "José", "minutos": 150},
        {"nome": "Ana", "minutos": 90},
    ]


def test_escreve_cabecalho_e_linhas() -> None:
    destino = io.BytesIO()
    exportador_xlsx.exportar(_linhas(), _COLUNAS, destino=destino)

    pasta = openpyxl.load_workbook(io.BytesIO(destino.getvalue()))
    aba = pasta[exportador_xlsx.NOME_ABA]
    linhas = list(aba.iter_rows(values_only=True))
    assert linhas[0] == ("Nome", "Horas")
    assert linhas[1] == ("José", 150)
    assert linhas[2] == ("Ana", 90)


def test_converte_apenas_coluna_marcada_como_duracao() -> None:
    destino = io.BytesIO()
    exportador_xlsx.exportar(_linhas(), _COLUNAS, destino=destino, converter_decimal=True)

    pasta = openpyxl.load_workbook(io.BytesIO(destino.getvalue()))
    aba = pasta[exportador_xlsx.NOME_ABA]
    linhas = list(aba.iter_rows(values_only=True))
    assert linhas[1] == ("José", 2.5)
    assert linhas[2] == ("Ana", 1.5)


def _gerador_10_mil_linhas() -> Iterator[dict[str, Any]]:
    for indice in range(10_000):
        yield {"nome": f"Colaborador {indice}", "minutos": indice}


def test_10_mil_linhas_via_gerador_nao_estoura_memoria() -> None:
    """Prova de streaming (PCF F11 T5, "pronto quando"): exportar 10.000
    linhas de um GERADOR (nunca uma lista materializada) usando
    `Workbook(write_only=True)` mantem o pico de alocacao rastreada muito
    abaixo do que custaria manter 10.000 linhas de celulas num workbook
    comum (`write_only=False`) -- se alguem remover `write_only=True` no
    futuro, este teste passa a falhar por estourar o teto."""
    destino = io.BytesIO()
    tracemalloc.start()
    try:
        exportador_xlsx.exportar(_gerador_10_mil_linhas(), _COLUNAS, destino=destino)
        _atual, pico = tracemalloc.get_traced_memory()
    finally:
        tracemalloc.stop()

    # Teto generoso (5 MB) para o pico de alocacao PYTHON rastreada durante a
    # exportacao (o buffer de saida em si, em `destino`, nao conta para
    # `tracemalloc` por ser um objeto ja existente antes do `start()`).
    # `Workbook(write_only=True)` grava cada linha no arquivo zip subjacente
    # conforme chega, sem reter celulas de linhas anteriores em memoria.
    assert pico < 5 * 1024 * 1024, f"pico de memoria rastreada: {pico} bytes"

    pasta = openpyxl.load_workbook(io.BytesIO(destino.getvalue()))
    aba = pasta[exportador_xlsx.NOME_ABA]
    assert aba.max_row == 10_001  # cabecalho + 10.000 linhas


def test_destino_permanece_aberto_apos_exportar() -> None:
    destino = io.BytesIO()
    exportador_xlsx.exportar(_linhas(), _COLUNAS, destino=destino)
    assert not destino.closed
