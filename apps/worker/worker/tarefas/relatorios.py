"""Execução assíncrona de relatórios (F11, T6 -- ownership conjunto de A1
"roteamento por dataset" + A3 "exportação/streaming/entrega", PCF §5).

Preenchido inteiro por A3: no momento em que esta tarefa foi escrita, A1
ainda não tinha chegado a T6 (`app/relatorios/motor.py`/`catalogo.py` já
publicados e estáveis, o que bastou para fixar o contrato usado aqui --
`app.relatorios.catalogo.obter_dataset`/`app.relatorios.motor.
ContextoConsulta` -- mas a integração literal dos dois lados de T6 não
tinha sido combinada ainda). Registrado como achado de coordenação no
relatório de fechamento da fase: se A1 preferir uma fronteira diferente
entre "resolver o dataset" e "exportar/entregar", é ajuste sobre este
arquivo, não reescrita -- os módulos de apoio (`catalogo`, `motor`,
`exportadores`, `entrega`) continuam exatamente os mesmos.

## Por que não usa `app.relatorios.motor.executar_dataset`

`motor.executar_dataset` pagina por deslocamento (`OFFSET`/`LIMIT`, teto
`motor.LIMITE_MAXIMO=5000`) -- desenhado para o caminho HTTP (síncrono e a
criação da execução assíncrona), não para os até ~372 mil linhas que esta
tarefa pode precisar exportar (PCF §2.3 item 3: "um relatório de 12 meses
por 1.000 colaboradores cai sempre" neste caminho). Em vez disso, esta
tarefa resolve o dataset da MESMA forma que o motor (`catalogo.
obter_dataset` + `motor.montar_contexto_consulta`), mas itera a `Select`
bruta devolvida pela função de dataset **em lotes por deslocamento
maiores** (`TAMANHO_LOTE`, não o teto pequeno do motor), escrevendo cada
lote no exportador conforme chega -- nunca materializa o resultado inteiro
em memória Python de uma vez (PCF §2.3 item 2). Uma única sessão/conexão
cobre lotes + atualização de progresso: `AsyncSession.execute(...)` sem
`stream()` não mantém cursor "preso" entre chamadas, então intercalar
`SELECT` de um lote com o `UPDATE` de progresso na mesma sessão é seguro
(nunca precisou de segunda conexão nem de ponte thread/fila para isto).

## Agrupamento

Esta tarefa **não suporta `agrupamento`** -- um relatório agrupado produz
poucas linhas por natureza (uma por grupo), então sempre cabe no caminho
síncrono ou na criação de execução assíncrona já materializada por
`app.relatorios.execucao._concluir_sincrono`/`_criar_assincrona` (A1, que
usa `motor.executar_dataset`, com suporte a `agrupamento`, para isso). Um
`parametros.agrupamento` presente aqui é ignorado deliberadamente (loga um
aviso) -- não é o caminho pesado que esta tarefa existe para cobrir.
"""

from __future__ import annotations

import datetime as dt
import hashlib
from typing import TYPE_CHECKING, Any
from uuid import UUID

from worker.filas import resultado_nao_implementado
from worker.log import obter_logger

if TYPE_CHECKING:
    from ponto_contracts import RelatorioDefinicao, RelatorioExecucao
    from sqlalchemy.ext.asyncio import AsyncSession

logger = obter_logger("tarefas.relatorios")

#: Linhas por lote de `OFFSET`/`LIMIT` -- grande o bastante para poucas
#: viagens ao banco em 372 mil linhas (~75 lotes), pequeno o bastante para
#: nunca reter mais que uma fração do resultado total em memória de uma vez.
TAMANHO_LOTE = 5000

#: Retenção do artefato -- mesmo valor de `app.relatorios.execucao.
#: RETENCAO_ARTEFATO` (A1); duplicado aqui pelo mesmo motivo documentado em
#: `app/core/filas.py` para `FILA_PADRAO` (apps/worker não importa apps/api
#: como pacote de aplicação, só como biblioteca instalada -- ADR-009 -- mas
#: uma constante pequena e estável não justifica acoplar os dois módulos).
RETENCAO_ARTEFATO = dt.timedelta(days=7)

_FORMATOS_SUPORTADOS = ("csv", "xlsx", "pdf")


class _EscritorCsv:
    """Escritor incremental de CSV -- mesma forma de saída de `app.
    relatorios.exportadores.csv` (BOM UTF-8, delimitador `,`), mas
    alimentado lote a lote em vez de um único iterável (ver docstring do
    módulo, "Por que não usa executar_dataset"). Duplica um pouco de lógica
    de baixo nível de `exportadores.csv` de propósito: aquele módulo
    continua com o contrato ONE-SHOT (`exportar(linhas, colunas, *,
    destino)`) que `app.relatorios.execucao` (A1, caminho síncrono, T3) já
    usa e testa; mudar sua assinatura para suportar chamadas incrementais
    quebraria esse contrato publicado. `app.relatorios.exportadores.decimal`
    (a única peça com REGRA de negócio, a fórmula de conversão) é
    reaproveitado sem duplicar."""

    content_type = "text/csv"

    def __init__(self) -> None:
        import csv
        import io

        self._buffer = io.StringIO()
        self._escritor = csv.writer(self._buffer)

    def cabecalho(self, colunas: list[Any]) -> None:
        self._escritor.writerow([c.rotulo for c in colunas])

    def linha(self, linha: dict[str, Any], colunas: list[Any], *, converter_decimal: bool) -> None:
        from app.relatorios.exportadores import valor_formatado  # type: ignore[import-not-found]

        self._escritor.writerow(
            [
                valor_formatado(linha, c, converter_decimal=converter_decimal, casas_decimais=2)
                for c in colunas
            ]
        )

    def finalizar(self) -> bytes:
        return b"\xef\xbb\xbf" + self._buffer.getvalue().encode("utf-8")


class _EscritorXlsx:
    """Escritor incremental de XLSX -- `openpyxl.Workbook(write_only=True)`
    já é, por natureza, um escritor incremental (cada `append` grava direto
    no zip subjacente); só falta o laço lote a lote, que esta classe
    fornece."""

    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __init__(self) -> None:
        import openpyxl

        self._pasta = openpyxl.Workbook(write_only=True)
        self._aba = self._pasta.create_sheet(title="Relatorio")

    def cabecalho(self, colunas: list[Any]) -> None:
        self._aba.append([c.rotulo for c in colunas])

    def linha(self, linha: dict[str, Any], colunas: list[Any], *, converter_decimal: bool) -> None:
        from app.relatorios.exportadores import valor_formatado

        self._aba.append(
            [
                valor_formatado(linha, c, converter_decimal=converter_decimal, casas_decimais=2)
                for c in colunas
            ]
        )

    def finalizar(self) -> bytes:
        import io

        buffer = io.BytesIO()
        self._pasta.save(buffer)
        return buffer.getvalue()


class _EscritorPdf:
    """Escritor incremental de PDF -- mesmo layout tabular genérico de
    `app.relatorios.exportadores.pdf` (cabeçalho/rodapé, zebra, paginação),
    reescrito para alimentação lote a lote. Relatório verdadeiramente
    gigante em PDF é impraticável por natureza do formato (não é o caminho
    pensado para 372 mil linhas -- CSV/XLSX são); esta classe existe para
    completude de formato, não para o caso extremo de volume."""

    content_type = "application/pdf"

    def __init__(self, *, titulo: str) -> None:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm

        self._colors = colors
        self._A4 = A4
        self._cm = cm
        self._margem = 1.6 * cm
        self._altura_linha = 0.5 * cm
        self._titulo = titulo
        self._colunas: list[Any] = []
        self._largura_coluna = 0.0
        self._pagina = 1
        self._indice_linha = 0
        self._y = 0.0

        import io

        from reportlab.pdfgen.canvas import Canvas

        self._buffer = io.BytesIO()
        self._canvas = Canvas(self._buffer, pagesize=A4)

    def _cabecalho_pagina(self) -> None:
        largura_pagina, altura_pagina = self._A4
        largura_util = largura_pagina - 2 * self._margem
        y = altura_pagina - self._margem
        if self._pagina == 1:
            self._canvas.setFont("Helvetica-Bold", 13)
            self._canvas.setFillColor(self._colors.HexColor("#262B31"))
            self._canvas.drawString(self._margem, y, self._titulo)
            y -= 0.7 * self._cm
        self._canvas.setFillColor(self._colors.HexColor("#4C5FCA"))
        self._canvas.rect(
            self._margem, y - self._altura_linha, largura_util, self._altura_linha, fill=1, stroke=0
        )
        self._canvas.setFillColor(self._colors.white)
        self._canvas.setFont("Helvetica-Bold", 7.5)
        for indice, coluna in enumerate(self._colunas):
            self._canvas.drawString(
                self._margem + indice * self._largura_coluna + 0.1 * self._cm,
                y - self._altura_linha + 0.14 * self._cm,
                coluna.rotulo[:40],
            )
        self._y = y - self._altura_linha

    def cabecalho(self, colunas: list[Any]) -> None:
        self._colunas = colunas
        largura_pagina, _ = self._A4
        largura_util = largura_pagina - 2 * self._margem
        self._largura_coluna = largura_util / max(len(colunas), 1)
        self._cabecalho_pagina()

    def linha(self, linha: dict[str, Any], colunas: list[Any], *, converter_decimal: bool) -> None:
        from app.relatorios.exportadores import valor_formatado

        largura_pagina, altura_pagina = self._A4
        if self._y - self._altura_linha < self._margem + 0.6 * self._cm:
            self._canvas.showPage()
            self._pagina += 1
            self._cabecalho_pagina()
        if self._indice_linha % 2 == 1:
            largura_util = largura_pagina - 2 * self._margem
            self._canvas.setFillColor(self._colors.HexColor("#EFF2F6"))
            self._canvas.rect(
                self._margem,
                self._y - self._altura_linha,
                largura_util,
                self._altura_linha,
                fill=1,
                stroke=0,
            )
        self._canvas.setFillColor(self._colors.HexColor("#262B31"))
        self._canvas.setFont("Courier", 7.5)
        for indice, coluna in enumerate(colunas):
            valor = valor_formatado(
                linha, coluna, converter_decimal=converter_decimal, casas_decimais=2
            )
            self._canvas.drawString(
                self._margem + indice * self._largura_coluna + 0.1 * self._cm,
                self._y - self._altura_linha + 0.14 * self._cm,
                str("" if valor is None else valor)[:40],
            )
        self._y -= self._altura_linha
        self._indice_linha += 1

    def finalizar(self) -> bytes:
        self._canvas.save()
        return self._buffer.getvalue()


_ESCRITORES: dict[str, Any] = {"csv": _EscritorCsv, "xlsx": _EscritorXlsx, "pdf": _EscritorPdf}


def _montar_contexto(tenant_id: UUID, parametros: dict[str, Any]) -> Any:
    """Reconstrói `ContextoConsulta` a partir do dicionário `parametros`
    (mesmas chaves camelCase que `app.relatorios.execucao.
    executar_relatorio_pedido` grava em `RelatorioExecucao.parametros` e
    passa ao enfileirar -- PCF, contrato entre as duas metades de T6)."""
    from app.relatorios.motor import montar_contexto_consulta  # type: ignore[import-not-found]

    def _uuid(valor: Any) -> UUID | None:
        return UUID(valor) if valor else None

    def _data(valor: Any) -> dt.date | None:
        return dt.date.fromisoformat(valor) if valor else None

    return montar_contexto_consulta(
        tenant_id,
        periodo_id=_uuid(parametros.get("periodoId")),
        de=_data(parametros.get("de")),
        ate=_data(parametros.get("ate")),
        empresa_id=_uuid(parametros.get("empresaId")),
        unidade_id=_uuid(parametros.get("unidadeId")),
        departamento_id=_uuid(parametros.get("departamentoId")),
        colaborador_id=_uuid(parametros.get("colaboradorId")),
        incluir_inativos=bool(parametros.get("incluirInativos")),
    )


async def _contar_total(sessao: AsyncSession, bruta: Any) -> int:
    import sqlalchemy as sa

    sub = bruta.subquery()
    resultado = await sessao.execute(sa.select(sa.func.count()).select_from(sub))
    return int(resultado.scalar_one())


async def _atualizar_progresso(
    sessao: AsyncSession, execucao: RelatorioExecucao, progresso: int
) -> None:
    """Grava o progresso e RE-PUBLICA `app.tenant_id` na sessao logo em
    seguida -- achado real (StaleDataError, ver `executar_relatorio`):
    `aplicar_tenant` usa `SET LOCAL` (escopo por transacao, `app/db/
    sessao.py`), e `commit()` encerra a transacao corrente. Sem republicar
    aqui, toda leitura/escrita seguinte desta MESMA sessao (o proximo lote
    do laco em `_executar`, ou os commits finais de `executar_relatorio`)
    roda sob RLS sem tenant nenhum -- falha fechada: SELECT devolve zero
    linhas silenciosamente, UPDATE por chave primaria não encontra a linha
    e o ORM levanta `StaleDataError` (nao ha `version_id_col` nesta tabela;
    o erro vem do proprio rowcount de UPDATE/DELETE que o SQLAlchemy sempre
    confere). Mesmo cuidado que `tests/f11/conftest.py::sessao_f11` ja
    documenta para o padrao de teste equivalente."""
    import app.db.sessao as _sessao_db  # type: ignore[import-not-found]

    execucao.progresso = max(0, min(100, progresso))
    await sessao.commit()
    await _sessao_db.aplicar_tenant(sessao, str(execucao.tenant_id))


async def _executar(
    sessao: AsyncSession,
    tenant_id: UUID,
    execucao: RelatorioExecucao,
    definicao: RelatorioDefinicao,
    *,
    parametros: dict[str, Any],
    formato: str,
) -> tuple[bytes, str, int]:
    """Roteamento por dataset (A1: `catalogo.obter_dataset`/`motor.
    ContextoConsulta`) + exportação em lotes/streaming (A3). Devolve
    `(conteudo, content_type, total_linhas)`."""
    # Importa por efeito colateral: e o que faz os 24 `registrar_dataset(...)`
    # de A2/A3/A4 (`app/relatorios/datasets/{operacionais,gerenciais,
    # espelho_oficial}.py`) rodarem antes da primeira chamada a
    # `obter_dataset` abaixo -- mesmo achado real que `app/routers/
    # relatorios.py` (A1) ja documenta e resolve para o processo da API,
    # mas o processo do WORKER nunca importa aquele roteador, entao precisa
    # do mesmo gatilho aqui (achado de integracao encontrado escrevendo o
    # teste desta tarefa, T6, registrado no relatorio de fechamento da fase).
    import app.relatorios.datasets as _datasets_registrados  # type: ignore[import-not-found] # noqa: F401
    import sqlalchemy as sa
    from app.relatorios.catalogo import (  # type: ignore[import-not-found]
        colunas_do_catalogo,
        obter_dataset,
    )

    if parametros.get("agrupamento"):
        logger.warning(
            "executar_relatorio: agrupamento pedido, ignorado no caminho assincrono de streaming",
            extra={"execucaoId": str(execucao.id), "agrupamento": parametros.get("agrupamento")},
        )

    contexto = _montar_contexto(tenant_id, parametros)
    construtor = obter_dataset(definicao.dataset)
    bruta = construtor(sessao, tenant_id, contexto)

    catalogo = colunas_do_catalogo(definicao)
    colunas_pedidas = parametros.get("colunas")
    if colunas_pedidas:
        por_chave = {c.chave: c for c in catalogo}
        colunas = [por_chave[chave] for chave in colunas_pedidas if chave in por_chave]
    else:
        colunas = list(catalogo)

    sub = bruta.subquery()
    projecao = [sub.c[c.chave] for c in colunas]
    consulta_base = sa.select(*projecao).select_from(sub)
    consulta_base = consulta_base.order_by(*[sa.asc(c) for c in consulta_base.selected_columns])

    total_estimado = await _contar_total(sessao, bruta)

    escritor_cls = _ESCRITORES[formato]
    escritor = escritor_cls(titulo=definicao.nome) if formato == "pdf" else escritor_cls()
    escritor.cabecalho(colunas)

    converter_decimal = bool(parametros.get("converterDecimal"))
    processadas = 0
    deslocamento = 0
    ultimo_progresso_gravado = -1
    while True:
        pagina = await sessao.execute(consulta_base.offset(deslocamento).limit(TAMANHO_LOTE))
        linhas_lote = pagina.mappings().all()
        if not linhas_lote:
            break
        for linha_bruta in linhas_lote:
            escritor.linha(dict(linha_bruta), colunas, converter_decimal=converter_decimal)
        processadas += len(linhas_lote)
        deslocamento += TAMANHO_LOTE

        progresso = int((processadas / total_estimado) * 100) if total_estimado else 99
        progresso = min(progresso, 99)  # 100 so no final, apos o upload confirmado
        if progresso != ultimo_progresso_gravado:
            await _atualizar_progresso(sessao, execucao, progresso)
            ultimo_progresso_gravado = progresso

        if len(linhas_lote) < TAMANHO_LOTE:
            break

    conteudo = escritor.finalizar()
    return conteudo, escritor.content_type, processadas


async def _entregar_agendamento(
    sessao: AsyncSession, execucao: RelatorioExecucao, agendamento_id: UUID
) -> None:
    """Quando `execucao.agendamento_id` não é nulo, chama a entrega do
    canal configurado (A3, T11) ao final -- `email.py`/`webhook.py`/
    `minio.py` (`app.relatorios.entrega`), interface `entregar(execucao,
    agendamento) -> bool` fixada pelo PCF §6/T11."""
    from ponto_contracts import RelatorioAgendamento as _RelatorioAgendamento

    agendamento = await sessao.get(_RelatorioAgendamento, agendamento_id)
    if agendamento is None:
        logger.warning(
            "executar_relatorio: agendamento nao encontrado para entrega",
            extra={"execucaoId": str(execucao.id), "agendamentoId": str(agendamento_id)},
        )
        return

    canais: dict[str, Any] = {}
    if agendamento.canal == "email":
        from app.relatorios.entrega import email as canal_email  # type: ignore[import-not-found]

        canais["email"] = canal_email.entregar
    elif agendamento.canal == "webhook":
        from app.relatorios.entrega import webhook as canal_webhook

        canais["webhook"] = canal_webhook.entregar
    else:
        from app.relatorios.entrega import minio as canal_minio

        canais["minio"] = canal_minio.entregar

    entregar = next(iter(canais.values()))
    sucesso = await entregar(execucao, agendamento)
    agendamento.ultima_execucao_em = dt.datetime.now(tz=dt.UTC)
    await sessao.commit()
    logger.info(
        "executar_relatorio: entrega do agendamento concluida",
        extra={
            "execucaoId": str(execucao.id),
            "agendamentoId": str(agendamento_id),
            "canal": agendamento.canal,
            "sucesso": sucesso,
        },
    )


async def executar_relatorio(
    ctx: dict[str, Any],
    *,
    tenant_id: str,
    execucao_id: str,
    codigo: str,
    parametros: dict[str, Any] | None = None,
    formato: str = "xlsx",
    solicitante_id: str | None = None,
) -> dict[str, Any]:
    """Executa um relatorio do catalogo e publica o arquivo no MinIO.

    Args:
        ctx: contexto do ARQ.
        tenant_id: tenant dono do dado.
        execucao_id: registro de execucao criado pela API, que recebe o progresso.
        codigo: codigo do relatorio no catalogo (ex.: `espelho-jornada`).
        parametros: periodo, filtros, agrupamentos e colunas escolhidas.
        formato: `csv`, `xlsx` ou `pdf`.
        solicitante_id: quem pediu, para a trilha de auditoria e para a entrega.
    """
    logger.info(
        "executar_relatorio recebida",
        extra={
            "jobId": ctx.get("job_id"),
            "tenantId": tenant_id,
            "execucaoId": execucao_id,
            "codigo": codigo,
            "formato": formato,
        },
    )

    if formato not in _FORMATOS_SUPORTADOS:
        return resultado_nao_implementado(
            "executar_relatorio",
            tenant_id=tenant_id,
            execucao_id=execucao_id,
            codigo=codigo,
            formato=formato,
        )

    # Import tardio de `app.*` (ADR-009) -- primeira ocorrencia textual do
    # modulo carrega o `# type: ignore`, mesmo padrao de `worker/tarefas/
    # fechamento.py`/`worker/notificacoes_verificacao.py`.
    from app.db.sessao import aplicar_tenant, fabrica_de_sessoes  # type: ignore[import-not-found]
    from ponto_contracts import RelatorioDefinicao as _RelatorioDefinicao
    from ponto_contracts import RelatorioExecucao as _RelatorioExecucao

    tenant_uuid = UUID(tenant_id)
    execucao_uuid = UUID(execucao_id)
    inicio = dt.datetime.now(tz=dt.UTC)

    fabrica = fabrica_de_sessoes()
    async with fabrica() as sessao:
        await aplicar_tenant(sessao, tenant_id)

        execucao = await sessao.get(_RelatorioExecucao, execucao_uuid)
        if execucao is None or execucao.tenant_id != tenant_uuid:
            logger.error(
                "executar_relatorio: execucao nao encontrada",
                extra={"jobId": ctx.get("job_id"), "execucaoId": execucao_id},
            )
            return {"implementado": True, "encontrado": False, "execucaoId": execucao_id}

        definicao = await sessao.get(_RelatorioDefinicao, execucao.relatorio_definicao_id)
        if definicao is None:
            execucao.status = "falhou"
            execucao.erro = "RelatorioDefinicao nao encontrada."
            await sessao.commit()
            return {"implementado": True, "sucesso": False, "execucaoId": execucao_id}

        execucao.status = "processando"
        execucao.progresso = 0
        await sessao.commit()
        # `commit()` acima encerra a transacao em que `aplicar_tenant` (SET
        # LOCAL, `app/db/sessao.py`) publicou `app.tenant_id` -- sem
        # republicar, `_executar` (que le o dataset sob RLS) e os commits de
        # falha/sucesso abaixo rodam sem tenant nenhum. Mesmo achado
        # documentado em `_atualizar_progresso` acima (achado real:
        # `StaleDataError` reproduzido em CI/local, ver git blame desta
        # linha).
        await aplicar_tenant(sessao, tenant_id)

        try:
            conteudo, content_type, total_linhas = await _executar(
                sessao,
                tenant_uuid,
                execucao,
                definicao,
                parametros=parametros or {},
                formato=formato,
            )
        except Exception as exc:
            logger.exception(
                "executar_relatorio: falha ao processar",
                extra={"jobId": ctx.get("job_id"), "execucaoId": execucao_id},
            )
            execucao.status = "falhou"
            execucao.erro = str(exc)
            await sessao.commit()
            return {
                "implementado": True,
                "sucesso": False,
                "execucaoId": execucao_id,
                "erro": str(exc),
            }

        hash_sha256 = hashlib.sha256(conteudo).hexdigest()
        chave = f"relatorios/{tenant_id}/{execucao_id}.{formato}"

        from app.comum.armazenamento import salvar_objeto  # type: ignore[import-not-found]

        try:
            await salvar_objeto(chave, conteudo, content_type=content_type)
        except Exception as exc:
            logger.exception(
                "executar_relatorio: falha ao gravar no armazenamento de objetos",
                extra={"jobId": ctx.get("job_id"), "execucaoId": execucao_id},
            )
            execucao.status = "falhou"
            execucao.erro = str(exc)
            await sessao.commit()
            return {
                "implementado": True,
                "sucesso": False,
                "execucaoId": execucao_id,
                "erro": str(exc),
            }

        concluido_em = dt.datetime.now(tz=dt.UTC)
        execucao.status = "concluido"
        execucao.progresso = 100
        execucao.total_linhas = total_linhas
        execucao.conteudo_ref = chave
        execucao.tamanho_bytes = len(conteudo)
        execucao.hash_sha256 = hash_sha256
        execucao.concluido_em = concluido_em
        execucao.duracao_ms = int((concluido_em - inicio).total_seconds() * 1000)
        execucao.expira_em = concluido_em + RETENCAO_ARTEFATO
        await sessao.commit()

        if execucao.agendamento_id is not None:
            # Mesmo achado dos outros pontos deste arquivo: `commit()` acima
            # encerra a transacao que publicava `app.tenant_id` -- sem
            # republicar, a leitura de `RelatorioAgendamento` e o commit
            # final dentro de `_entregar_agendamento` rodam sem tenant.
            await aplicar_tenant(sessao, tenant_id)
            await _entregar_agendamento(sessao, execucao, execucao.agendamento_id)

        logger.info(
            "executar_relatorio concluida",
            extra={
                "jobId": ctx.get("job_id"),
                "execucaoId": execucao_id,
                "totalLinhas": total_linhas,
                "duracaoMs": execucao.duracao_ms,
            },
        )
        return {
            "implementado": True,
            "sucesso": True,
            "execucaoId": execucao_id,
            "totalLinhas": total_linhas,
            "hashSha256": hash_sha256,
        }
